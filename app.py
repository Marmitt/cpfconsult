'''
1 - Open excel and extract cpf of the client
2 - Open the website: https://consultcpf-devaprender.netlify.app/ and use excel to check the status of the payment
3 - Check if it is "alright" or "late"
4 - If it is alright, get the date and the method of payment
5 - If it is late, mark as "pendent"
6 - Insert all the info (name, value, cpf, expire-date, status, date of the payment, method of payment (card, invoice) in another excel
7 - Repeat until the last client
'''
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# 1 - Open excel and extract cpf of the client
clients_data = openpyxl.load_workbook('clients_data.xlsx')
clients_page = clients_data['Sheet1']

driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')

for line in clients_page.iter_rows(min_row=2, values_only=True):
    name, value, cpf, expire_date = line

# 2 - Open the website: https://consultcpf-devaprender.netlify.app/ and use excel to check the status of the payment
    sleep(5)
    search = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    search.clear()
    search.send_keys(cpf)
    sleep(1)

    # 3 - Check if it is "alright" or "late"
    button_search = driver.find_element(
        By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    button_search.click()
    sleep(4)
    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")

    if status.text == 'em dia':
        # 4 - If it is alright, get the date and the method of payment
        payment_date = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        method_payment = driver.find_element(
            By.XPATH, "//p[@id='paymentMethod']")

        payment_date_clear = payment_date.text.split()[3]
        method_payment_clear = method_payment.text.split()[3]

        updated_data = openpyxl.load_workbook('updated_data.xlsx')
        updated_page = updated_data['Sheet1']

        updated_page.append(
            [name, value, cpf, expire_date, 'em dia', payment_date_clear, method_payment_clear])

        updated_data.save('updated_data.xlsx')
    else:
        # 5 - If it is late, mark as "pending"
        updated_data = openpyxl.load_workbook('updated_data.xlsx')
        updated_page = updated_data['Sheet1']

        updated_page.append([name, value, cpf, expire_date, 'pending'])

        updated_data.save('updated_data.xlsx')
